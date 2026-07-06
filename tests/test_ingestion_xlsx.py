"""Unit tests for xlsx ingestion (S-11).

Builds a minimal workbook in memory with openpyxl,
then verifies that read_xlsx() extracts sheet names, columns, and row counts.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.ingestion.xlsx import read_xlsx


@pytest.fixture()
def single_sheet_xlsx(tmp_path: Path) -> Path:
    """One sheet with 3 columns and 2 data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Product", "Units", "Revenue"])
    ws.append(["Widget A", 100, 2500.0])
    ws.append(["Widget B", 200, 4800.0])
    p = tmp_path / "sales.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture()
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Two sheets: Summary and Details."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Category", "Total"])
    ws1.append(["Alpha", 100])

    ws2 = wb.create_sheet("Details")
    ws2.append(["Item", "Qty", "Price", "Note"])
    ws2.append(["X", 5, 9.99, "first"])
    ws2.append(["Y", 3, 14.99, "second"])
    ws2.append(["Z", 7, 4.99, "third"])
    p = tmp_path / "multi.xlsx"
    wb.save(str(p))
    return p


class TestReadXlsxSingleSheet:
    def test_sheet_name_in_text(self, single_sheet_xlsx):
        text, _ = read_xlsx(single_sheet_xlsx)
        assert "Sales" in text

    def test_column_headers_in_text(self, single_sheet_xlsx):
        text, _ = read_xlsx(single_sheet_xlsx)
        assert "Product" in text
        assert "Units" in text
        assert "Revenue" in text

    def test_data_in_text(self, single_sheet_xlsx):
        text, _ = read_xlsx(single_sheet_xlsx)
        assert "Widget A" in text

    def test_structured_meta_sheet_name(self, single_sheet_xlsx):
        _, meta = read_xlsx(single_sheet_xlsx)
        assert "Sales" in meta["sheets"]

    def test_structured_meta_row_count(self, single_sheet_xlsx):
        _, meta = read_xlsx(single_sheet_xlsx)
        assert meta["sheets"]["Sales"]["row_count"] == 2

    def test_structured_meta_columns(self, single_sheet_xlsx):
        _, meta = read_xlsx(single_sheet_xlsx)
        cols = meta["sheets"]["Sales"]["columns"]
        assert "Product" in cols
        assert "Units" in cols
        assert "Revenue" in cols


class TestReadXlsxMultiSheet:
    def test_both_sheet_names_present(self, multi_sheet_xlsx):
        text, meta = read_xlsx(multi_sheet_xlsx)
        assert "Summary" in meta["sheets"]
        assert "Details" in meta["sheets"]

    def test_summary_row_count(self, multi_sheet_xlsx):
        _, meta = read_xlsx(multi_sheet_xlsx)
        assert meta["sheets"]["Summary"]["row_count"] == 1

    def test_details_row_count(self, multi_sheet_xlsx):
        _, meta = read_xlsx(multi_sheet_xlsx)
        assert meta["sheets"]["Details"]["row_count"] == 3


class TestReadXlsxErrorHandling:
    def test_nonexistent_file_returns_error_string(self, tmp_path):
        bad = tmp_path / "missing.xlsx"
        text, meta = read_xlsx(bad)
        assert "error" in text.lower()
        assert meta == {}
