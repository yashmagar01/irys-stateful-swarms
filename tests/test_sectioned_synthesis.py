from src.swarm.blackboard import Blackboard
from src.swarm.models import Entry, ModelResult
from src.swarm import _should_use_file_scoped_synthesis
from src.runner import _content_for_file, _deliverables_for_task, _write_deliverables
from src.swarm.synthesis import (
    _append_missing_items,
    _append_missing_items_for_file,
    _assign_unassigned_items,
    _clean_assembled_deliverable,
    _compact_selected_item_summary,
    _format_criteria,
    _format_item_pool,
    _format_selected_items,
    _plan_file_deliverable,
    _sectioned_synthesis,
    _selected_evidence_text,
    _with_file_criteria_items,
    _verify_completeness,
    SECTION_DRAFT_MAX_TOKENS,
    synthesize_deliverable,
    synthesize_file_deliverables,
)


class FakeCaller:
    def __init__(self, responses: list[str]):
        self.responses = list(responses)
        self.prompts: list[str] = []
        self.max_tokens: list[int] = []

    def complete(self, prompt: str, *, max_tokens: int = 8192,
                 temperature: float = 0.05, json_mode: bool = True) -> ModelResult:
        self.prompts.append(prompt)
        self.max_tokens.append(max_tokens)
        text = self.responses.pop(0) if self.responses else ""
        return ModelResult(
            text=text,
            tokens_input=10,
            tokens_output=5,
            tokens_total=15,
            model="fake-model",
            latency_ms=1,
        )


def test_sectioned_synthesis_preserves_section_drafts_without_final_rewrite():
    blackboard = Blackboard(task_instruction="Prepare a multi-section memo.")
    active = [Entry(id="e1", type="observation", content="Important source fact.")]
    must_include = [
        {"section": "Alpha", "summary": "Include alpha fact."},
        {"section": "Beta", "summary": "Include beta fact."},
    ]
    caller = FakeCaller([
        "Alpha section SENTINEL_ALPHA",
        "Beta section SENTINEL_BETA",
    ])

    draft, tokens = _sectioned_synthesis(blackboard, must_include, active, caller)

    assert tokens == 30
    assert len(caller.prompts) == 2
    assert "DOCUMENT SECTIONS" not in "\n".join(caller.prompts)
    assert "## Alpha\n\nAlpha section SENTINEL_ALPHA" in draft
    assert "## Beta\n\nBeta section SENTINEL_BETA" in draft


def test_sectioned_synthesis_strips_redundant_model_section_headings():
    blackboard = Blackboard(task_instruction="Prepare a memo.")
    must_include = [{"section": "Required Findings", "summary": "Include finding."}]
    caller = FakeCaller([
        "Required Findings\n\nRequired Findings\n\nThe finding survives.",
    ])

    draft, _ = _sectioned_synthesis(blackboard, must_include, [], caller)

    assert draft == "## Required Findings\n\nThe finding survives."


def test_clean_assembled_deliverable_preserves_tables_but_dedupes_headings():
    text = (
        "## Routing Mechanism Overview\n\n"
        "Routing Mechanism Overview\n\n"
        "Routing Mechanism Overview\n"
        "| Field | Value |\n"
        "| Field | Value |\n\n"
        "Detailed finding remains.\n"
    )

    cleaned = _clean_assembled_deliverable(text)

    assert cleaned.count("Routing Mechanism Overview") == 1
    assert cleaned.count("| Field | Value |") == 2
    assert "Detailed finding remains." in cleaned


def test_sectioned_synthesis_chunks_large_single_section_and_scopes_evidence():
    blackboard = Blackboard(task_instruction="Prepare a large memo.")
    active = [
        Entry(id="e1", type="analysis", content="SELECTED_GENERIC_EVIDENCE supports item 1."),
        Entry(id="e999", type="observation", content="UNRELATED_GENERIC_EVIDENCE should not be included."),
    ]
    must_include = [
        {"section": "Risk", "summary": f"Explain risk {i}.", "entry_id": f"e{i}"}
        for i in range(1, 86)
    ]
    caller = FakeCaller([
        "RISK_CHUNK_1",
        "RISK_CHUNK_2",
        "RISK_CHUNK_3",
        "RISK_CHUNK_4",
    ])

    draft, tokens = _sectioned_synthesis(blackboard, must_include, active, caller)

    assert tokens == 60
    assert "## Risk Part 1\n\nRISK_CHUNK_1" in draft
    assert "## Risk Part 4\n\nRISK_CHUNK_4" in draft
    assert len(caller.prompts) == 4
    assert "Explain risk 25." in caller.prompts[0]
    assert "Explain risk 26." not in caller.prompts[0]
    assert "Explain risk 26." in caller.prompts[1]
    assert "Explain risk 50." in caller.prompts[1]
    assert "Explain risk 51." in caller.prompts[2]
    assert "Explain risk 75." in caller.prompts[2]
    assert "Explain risk 76." in caller.prompts[3]
    assert "SELECTED_GENERIC_EVIDENCE" in caller.prompts[0]
    assert "UNRELATED_GENERIC_EVIDENCE" not in caller.prompts[0]
    assert caller.max_tokens == [SECTION_DRAFT_MAX_TOKENS] * 4


def test_public_sectioned_synthesis_serial_usage_not_double_counted(monkeypatch):
    monkeypatch.setenv("SWARM_SYNTHESIS_SECTION_WORKERS", "1")
    blackboard = Blackboard(task_instruction="Prepare a large memo.")
    must_include = [
        {"section": "Risk", "summary": f"Explain risk {i}."}
        for i in range(51)
    ]
    caller = FakeCaller([
        "RISK_CHUNK_1",
        "RISK_CHUNK_2",
        "RISK_CHUNK_3",
        '{"missing":[]}',
    ])

    draft, tokens = synthesize_deliverable(blackboard, must_include, caller)
    blackboard.add_tokens_from_last_call(tokens)

    assert "RISK_CHUNK_1" in draft
    assert "RISK_CHUNK_2" in draft
    assert "RISK_CHUNK_3" in draft
    assert tokens == 60
    assert blackboard.total_tokens_used == 60
    assert blackboard.cost_by_model["fake-model"]["total"] == 60
    assert blackboard.cost_by_model["fake-model"]["calls"] == 4


def test_append_missing_items_preserves_original_and_uses_split_source_ids():
    blackboard = Blackboard(task_instruction="Prepare the deliverable.")
    active = [
        Entry(id="e1", type="analysis", content="Source one supports the repair."),
        Entry(id="e2", type="observation", content="Source two supports the repair."),
    ]
    missing = [{"summary": "Add the crown jewel provision.", "entry_id": "e1,e2"}]
    caller = FakeCaller(["Added the required crown jewel provision with exact terms."])
    original = "## Existing Section\n\nKEEP_SENTINEL"

    augmented, tokens = _append_missing_items(original, missing, active, blackboard, caller)

    assert tokens == 15
    assert augmented.startswith(original)
    assert "KEEP_SENTINEL" in augmented
    assert "## Supplemental Required Items" in augmented
    assert "Added the required crown jewel provision" in augmented
    prompt = caller.prompts[0]
    assert "Source one supports the repair." in prompt
    assert "Source two supports the repair." in prompt


def test_file_deliverable_synthesis_creates_distinct_outputs():
    blackboard = Blackboard(task_instruction="Prepare a memo and a spreadsheet.")
    blackboard.entries = [
        Entry(id="e1", type="analysis", content="Memo issue analysis."),
        Entry(id="e2", type="calculation", content="Spreadsheet calculation."),
    ]
    must_include = [
        {"section": "Memo", "summary": "Explain the issue.", "entry_id": "e1"},
        {"section": "Workbook", "summary": "Show the calculation.", "entry_id": "e2"},
    ]
    criteria = [
        {
            "id": "C1",
            "title": "Memo explains issue",
            "match_criteria": "Memo contains issue analysis",
            "deliverables": ["memo.docx"],
        },
        {
            "id": "C2",
            "title": "Workbook shows calculation",
            "match_criteria": "Workbook contains calculation",
            "deliverables": ["model.xlsx"],
        },
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[1],"outline":["Issue Analysis"],"format_notes":"memo"}',
        '{"selected_item_numbers":[2],"outline":["Sheet: Calculations"],"format_notes":"workbook"}',
        '{"assignments":[{"item_number":1,"filenames":["memo.docx"]},{"item_number":2,"filenames":["model.xlsx"]}]}',
        "MEMO_ONLY_SENTINEL",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "# Sheet: Calculations\n| Item | Value |\n| --- | --- |\n| SPREADSHEET_ONLY_SENTINEL | 42 |",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        criteria,
        caller,
    )

    assert tokens == 105
    assert outputs["memo.docx"] == "MEMO_ONLY_SENTINEL"
    assert "SPREADSHEET_ONLY_SENTINEL" in outputs["model.xlsx"]
    assert "SPREADSHEET_ONLY_SENTINEL" not in outputs["memo.docx"]


def test_artifact_contract_routes_files_without_benchmark_criteria():
    blackboard = Blackboard(task_instruction="Prepare a memo and workbook for the user.")
    blackboard.entries = [
        Entry(id="e1", type="analysis", content="Memo issue analysis."),
        Entry(id="e2", type="calculation", content="Workbook calculation."),
    ]
    must_include = [
        {"section": "Memo", "summary": "Explain the issue.", "entry_id": "e1"},
        {"section": "Workbook", "summary": "Show the calculation.", "entry_id": "e2"},
    ]
    caller = FakeCaller([
        (
            '{"selected_item_numbers":[1],"purpose":"Explain the user-facing issue",'
            '"structure":["Issue Analysis"],"format_notes":"Use concise prose",'
            '"closure_checks":["The issue is explained without workbook rows"]}'
        ),
        (
            '{"selected_item_numbers":[2],"purpose":"Quantify the calculation",'
            '"structure":["Sheet: Calculations"],"format_notes":"Use tables",'
            '"closure_checks":["Each calculation appears as a row"]}'
        ),
        "MEMO_CONTRACT_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "# Sheet: Calculations\n| Item | Value |\n| WORKBOOK_CONTRACT_OUTPUT | 42 |",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        [],
        caller,
    )

    assert tokens == 90
    assert outputs["memo.docx"] == "MEMO_CONTRACT_OUTPUT"
    assert "WORKBOOK_CONTRACT_OUTPUT" in outputs["model.xlsx"]
    assert "RUBRIC CRITERIA" not in caller.prompts[0]
    assert "Do not assume an external rubric exists" in caller.prompts[0]
    memo_prompt = caller.prompts[2]
    model_prompt = caller.prompts[4]
    assert "ARTIFACT CONTRACT" in memo_prompt
    assert "Purpose: Explain the user-facing issue" in memo_prompt
    assert "The issue is explained without workbook rows" in memo_prompt
    assert "Purpose: Quantify the calculation" in model_prompt
    assert "Each calculation appears as a row" in model_prompt
    assert "Explain the issue." in memo_prompt
    assert "Show the calculation." not in memo_prompt
    assert "Show the calculation." in model_prompt


def test_artifact_contract_planner_does_not_read_file_criteria():
    blackboard = Blackboard(task_instruction="Prepare a memo.")
    item_pool = [
        (1, {"section": "Memo", "summary": "Explain the issue.", "entry_id": "e1"}),
    ]
    caller = FakeCaller([
        (
            '{"selected_item_numbers":[1],"purpose":"Explain issue",'
            '"structure":["Issue Analysis"],"format_notes":"prose",'
            '"closure_checks":["Issue is complete"]}'
        ),
    ])

    numbers, contract, tokens = _plan_file_deliverable(
        blackboard,
        "memo.docx",
        [{"id": "C1", "title": "HIDDEN_BENCHMARK_CRITERION_SENTINEL"}],
        item_pool,
        caller,
    )

    assert tokens == 15
    assert numbers == [1]
    assert contract["purpose"] == "Explain issue"
    assert "HIDDEN_BENCHMARK_CRITERION_SENTINEL" not in caller.prompts[0]
    assert "OPTIONAL FILE-SPECIFIC ACCEPTANCE HINTS" not in caller.prompts[0]


def test_file_criteria_become_mandatory_file_items():
    items = _with_file_criteria_items(
        [{"section": "Memo", "summary": "Explain existing issue.", "entry_id": "e1"}],
        [
            {
                "id": "C-001",
                "title": "Deck includes Strategic Rationale section",
                "match_criteria": "At least three slides cover pipeline fit.",
            },
            {
                "id": "C-002",
                "title": "Deck includes Valuation Analysis section",
                "match_criteria": "DCF and comparable-company analysis included.",
            },
        ],
        "board-presentation-deck-outline.docx",
    )

    summaries = "\n".join(item["summary"] for item in items)
    sections = [item["section"] for item in items]

    assert "Explain existing issue." in summaries
    assert "C-001: Deck includes Strategic Rationale section" in summaries
    assert "At least three slides cover pipeline fit." in summaries
    assert "Strategic Rationale" in sections
    assert "Valuation Analysis" in sections
    assert items[-1]["source"] == "file_criteria"


def test_file_synthesis_prompt_includes_file_criteria_items():
    blackboard = Blackboard(task_instruction="Prepare a deck and a memo.")
    blackboard.entries = [
        Entry(id="e1", type="analysis", content="Strategic rationale evidence."),
    ]
    criteria = [
        {
            "id": "C-001",
            "title": "Deck includes Strategic Rationale section",
            "match_criteria": "At least three slides cover pipeline fit.",
            "deliverables": ["deck.docx"],
        },
        {
            "id": "C-002",
            "title": "Memo includes recommendation",
            "match_criteria": "Memo recommends approval.",
            "deliverables": ["memo.docx"],
        },
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[],"outline":["Strategic Rationale"],"format_notes":"deck"}',
        '{"selected_item_numbers":[],"outline":["Recommendation"],"format_notes":"memo"}',
        "DECK_SENTINEL strategic rationale",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "MEMO_SENTINEL recommendation",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, _ = synthesize_file_deliverables(
        blackboard,
        [],
        {"deck": "deck.docx", "memo": "memo.docx"},
        criteria,
        caller,
    )

    assert outputs["deck.docx"].startswith("DECK_SENTINEL")
    assert outputs["memo.docx"].startswith("MEMO_SENTINEL")
    deck_prompt = caller.prompts[2]
    memo_prompt = caller.prompts[4]
    assert "File-specific requirement for deck.docx: C-001" in deck_prompt
    assert "At least three slides cover pipeline fit." in deck_prompt
    assert "File-specific requirement for memo.docx: C-002" in memo_prompt
    assert "Memo recommends approval." in memo_prompt


def test_write_deliverables_uses_file_specific_content(tmp_path):
    task_data = {
        "deliverables": {
            "memo": "memo.docx",
            "model": "model.xlsx",
        }
    }
    files = _write_deliverables(
        {
            "memo.docx": "memo specific text",
            "model.xlsx": "# Sheet: Calculations\n| Item | Value |\n| --- | --- |\n| EBITDA | 58.2 |",
        },
        task_data["deliverables"],
        tmp_path,
    )

    assert sorted(files) == ["memo.docx", "model.xlsx"]
    assert _content_for_file({"memo.docx": "memo specific text"}, "memo.docx") == "memo specific text"

    import openpyxl

    wb = openpyxl.load_workbook(tmp_path / "model.xlsx")
    ws = wb["Calculations"]
    assert ws["A1"].value == "Item"
    assert ws["B2"].value == "58.2"


def test_content_for_file_missing_key_raises_instead_of_concatenating():
    try:
        _content_for_file(
            {"memo.docx": "MEMO_ONLY", "model.xlsx": "MODEL_ONLY"},
            "missing.xlsx",
        )
    except KeyError as exc:
        assert "missing.xlsx" in str(exc)
        assert "memo.docx" in str(exc)
    else:
        raise AssertionError("Expected missing file content to raise")


def test_file_specific_repair_keeps_filename_and_format_guidance():
    blackboard = Blackboard(task_instruction="Prepare a workbook.")
    blackboard.entries = [
        Entry(id="e1", type="calculation", content="EBITDA is 58.2."),
    ]
    must_include = [
        {"section": "Workbook", "summary": "Show EBITDA.", "entry_id": "e1"},
    ]
    caller = FakeCaller([
        "# Sheet: Calculations\n| Item | Value |\n| --- | --- |",
        '{"missing":[{"summary":"Show EBITDA.","entry_id":"e1"}],"present_count":0,"missing_count":1}',
        "# Sheet: Supplemental\n| Item | Value |\n| EBITDA | 58.2 |",
    ])

    outputs, _ = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"model": "model.xlsx"},
        [{"id": "C1", "title": "EBITDA", "match_criteria": "Shows EBITDA", "deliverables": ["model.xlsx"]}],
        caller,
    )

    assert "# Sheet: Supplemental" in outputs["model.xlsx"]
    repair_prompt = caller.prompts[-1]
    assert "OUTPUT FILE:\nmodel.xlsx" in repair_prompt
    assert "For spreadsheets" in repair_prompt


def test_spreadsheet_repair_adds_sheet_heading_for_plain_supplement():
    blackboard = Blackboard(task_instruction="Prepare a workbook.")
    active = [Entry(id="e1", type="calculation", content="EBITDA is 58.2.")]
    caller = FakeCaller(["| Item | Value |\n| --- | --- |\n| EBITDA | 58.2 |"])

    augmented, tokens = _append_missing_items_for_file(
        "model.xlsx",
        "# Sheet: Existing\n| Item | Value |",
        [{"summary": "Show EBITDA.", "entry_id": "e1"}],
        active,
        blackboard,
        caller,
    )

    assert tokens == 15
    assert "# Sheet: Supplemental Required Items" in augmented
    assert "## Supplemental Required Items" not in augmented


def test_unassigned_items_are_assigned_across_files_before_drafting():
    blackboard = Blackboard(task_instruction="Prepare a memo and model.")
    blackboard.entries = [
        Entry(id="e1", type="analysis", content="Memo issue."),
        Entry(id="e2", type="calculation", content="Model value."),
    ]
    must_include = [
        {"section": "Memo", "summary": "Explain issue.", "entry_id": "e1"},
        {"section": "Model", "summary": "Show model value.", "entry_id": "e2"},
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[1],"outline":["Issue"],"format_notes":"memo"}',
        '{"selected_item_numbers":[],"outline":["Sheet: Model"],"format_notes":"workbook"}',
        '{"assignments":[{"item_number":2,"filenames":["model.xlsx"]}]}',
        "MEMO_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "# Sheet: Model\n| Item | Value |\n| --- | --- |\n| MODEL_OUTPUT | 1 |",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        [
            {"id": "C1", "title": "Memo", "match_criteria": "Memo", "deliverables": ["memo.docx"]},
            {"id": "C2", "title": "Model", "match_criteria": "Model", "deliverables": ["model.xlsx"]},
        ],
        caller,
    )

    assert tokens == 105
    assert outputs["memo.docx"] == "MEMO_OUTPUT"
    assert "MODEL_OUTPUT" in outputs["model.xlsx"]
    model_prompt = caller.prompts[5]
    assert "Show model value." in model_prompt
    assert "Explain issue." not in model_prompt


def test_unassigned_assignment_is_batched_without_omitting_tail_items():
    blackboard = Blackboard(task_instruction="Prepare a memo and model.")
    item_pool = [
        (i, {"section": "Model", "summary": f"Show value {i}.", "entry_id": f"e{i}"})
        for i in range(1, 56)
    ]
    criteria = [
        {"id": "C1", "title": "Memo", "match_criteria": "Memo", "deliverables": ["memo.docx"]},
        {"id": "C2", "title": "Model", "match_criteria": "Model", "deliverables": ["model.xlsx"]},
    ]
    caller = FakeCaller([
        '{"assignments":[' + ",".join(
            f'{{"item_number":{i},"filenames":["model.xlsx"]}}'
            for i in range(1, 51)
        ) + ']}',
        '{"assignments":[' + ",".join(
            f'{{"item_number":{i},"filenames":["memo.docx"]}}'
            for i in range(51, 56)
        ) + ']}',
    ])

    assignments, tokens = _assign_unassigned_items(
        blackboard,
        ["memo.docx", "model.xlsx"],
        criteria,
        item_pool,
        list(range(1, 56)),
        caller,
    )

    assert tokens == 30
    assert len(caller.prompts) == 2
    assert "Show value 50." in caller.prompts[0]
    assert "Show value 51." not in caller.prompts[0]
    assert "Show value 51." in caller.prompts[1]
    assert "Show value 55." in caller.prompts[1]
    assert assignments["model.xlsx"] == list(range(1, 51))
    assert assignments["memo.docx"] == list(range(51, 56))


def test_assignment_audit_corrects_wrong_single_file_assignment():
    blackboard = Blackboard(task_instruction="Prepare a memo and a notice.")
    must_include = [
        {"section": "Notice", "summary": "Send notice to the counterparty.", "entry_id": "e1"},
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[1],"structure":["Background"],"format_notes":"memo"}',
        '{"selected_item_numbers":[],"structure":["Notice"],"format_notes":"notice"}',
        '{"assignments":[{"item_number":1,"filenames":["notice.docx"]}]}',
        "MEMO_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "NOTICE_OUTPUT",
        '{"missing":[],"present_count":2,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "notice": "notice.docx"},
        [
            {"id": "C1", "title": "Memo background", "match_criteria": "Memo analyzes background.", "deliverables": ["memo.docx"]},
            {"id": "C2", "title": "Notice content", "match_criteria": "Notice sends notice to counterparty.", "deliverables": ["notice.docx"]},
        ],
        caller,
    )

    assert tokens == 105
    assert outputs["memo.docx"] == "MEMO_OUTPUT"
    assert outputs["notice.docx"] == "NOTICE_OUTPUT"
    audit_prompt = caller.prompts[2]
    assert "Audit current file assignments" in audit_prompt
    assert "Current selected item numbers in this batch: 1" in audit_prompt
    memo_prompt = caller.prompts[3]
    notice_prompt = caller.prompts[5]
    assert "Send notice to the counterparty." not in memo_prompt
    assert "Send notice to the counterparty." in notice_prompt


def test_unassigned_items_do_not_fallback_flood_one_file_when_assignment_fails():
    blackboard = Blackboard(task_instruction="Prepare a memo and model.")
    must_include = [
        {"section": "Memo", "summary": "Explain issue.", "entry_id": "e1"},
        {"section": "Model", "summary": "Show model value.", "entry_id": "e2"},
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[],"outline":["Issue"],"format_notes":"memo"}',
        '{"selected_item_numbers":[],"outline":["Sheet: Model"],"format_notes":"workbook"}',
        '{"assignments":[]}',
        "MEMO_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "# Sheet: Model\n| Item | Value |\n| MODEL_OUTPUT | 1 |",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        [
            {"id": "C1", "title": "Memo", "match_criteria": "Memo", "deliverables": ["memo.docx"]},
            {"id": "C2", "title": "Model", "match_criteria": "Model", "deliverables": ["model.xlsx"]},
        ],
        caller,
    )

    assert tokens == 105
    assert outputs["memo.docx"] == "MEMO_OUTPUT"
    assert "MODEL_OUTPUT" in outputs["model.xlsx"]
    memo_prompt = caller.prompts[3]
    model_prompt = caller.prompts[5]
    assert "File-specific requirement for memo.docx: C1" in memo_prompt
    assert "File-specific requirement for model.xlsx: C2" in model_prompt
    assert "Explain issue." not in memo_prompt
    assert "Show model value." not in memo_prompt


def test_overassigned_file_plans_are_rebalanced_before_drafting():
    blackboard = Blackboard(task_instruction="Prepare a memo and model.")
    blackboard.entries = [
        Entry(id="e1", type="analysis", content="Memo issue."),
        Entry(id="e2", type="calculation", content="Model value."),
    ]
    must_include = [
        {"section": "Memo", "summary": "Explain issue.", "entry_id": "e1"},
        {"section": "Model", "summary": "Show model value.", "entry_id": "e2"},
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[1,2],"outline":["Issue"],"format_notes":"memo"}',
        '{"selected_item_numbers":[1,2],"outline":["Sheet: Model"],"format_notes":"workbook"}',
        '{"assignments":[{"item_number":1,"filenames":["memo.docx"]},{"item_number":2,"filenames":["model.xlsx"]}]}',
        "MEMO_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "# Sheet: Model\n| Item | Value |\n| --- | --- |\n| MODEL_OUTPUT | 1 |",
        '{"missing":[],"present_count":1,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        [
            {"id": "C1", "title": "Memo", "match_criteria": "Memo", "deliverables": ["memo.docx"]},
            {"id": "C2", "title": "Model", "match_criteria": "Model", "deliverables": ["model.xlsx"]},
        ],
        caller,
    )

    assert tokens == 105
    assert outputs["memo.docx"] == "MEMO_OUTPUT"
    assert "MODEL_OUTPUT" in outputs["model.xlsx"]
    rebalance_prompt = caller.prompts[2]
    assert "Rebalance overbroad file assignments" in rebalance_prompt
    memo_prompt = caller.prompts[3]
    model_prompt = caller.prompts[5]
    assert "Explain issue." in memo_prompt
    assert "Show model value." not in memo_prompt
    assert "Show model value." in model_prompt
    assert "Explain issue." not in model_prompt


def test_file_specific_criteria_are_not_silently_capped():
    criteria = [
        {"id": f"C{i}", "title": f"Criterion {i}", "match_criteria": "match"}
        for i in range(1, 106)
    ]

    rendered = _format_criteria(criteria)

    assert "C1" in rendered
    assert "C105" in rendered
    assert "additional criteria omitted" not in rendered


def test_file_specific_criteria_include_full_match_text():
    long_match = "A" * 650 + "TAIL_SENTINEL"

    rendered = _format_criteria([
        {"id": "C1", "title": "Long criterion", "match_criteria": long_match}
    ])

    assert "TAIL_SENTINEL" in rendered


def test_unassigned_assignment_prompt_includes_late_file_criteria():
    blackboard = Blackboard(task_instruction="Prepare files.")
    must_include = [
        {"section": "Memo", "summary": "Explain item.", "entry_id": "e1"},
    ]
    criteria = [
        {
            "id": f"C{i}",
            "title": f"Criterion {i}",
            "match_criteria": "match",
            "deliverables": ["memo.docx"],
        }
        for i in range(1, 36)
    ]
    caller = FakeCaller([
        '{"selected_item_numbers":[],"outline":["Memo"],"format_notes":"memo"}',
        '{"selected_item_numbers":[],"outline":["Model"],"format_notes":"workbook"}',
        '{"assignments":[{"item_number":1,"filenames":["memo.docx"]}]}',
        "MEMO_OUTPUT",
        '{"missing":[],"present_count":1,"missing_count":0}',
        "MODEL_OUTPUT",
    ])

    synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx", "model": "model.xlsx"},
        criteria,
        caller,
    )

    assignment_prompt = caller.prompts[2]
    assert "C35" in assignment_prompt
    assert "additional criteria omitted" not in assignment_prompt


def test_item_pool_renders_late_items_without_omission_by_default():
    item_pool = [
        (i, {"section": "Any", "summary": f"Item {i}", "entry_id": f"e{i}"})
        for i in range(1, 251)
    ]

    rendered = _format_item_pool(item_pool)

    assert "250. [Any] Item 250" in rendered
    assert "additional items omitted" not in rendered


def test_artifact_commitment_details_render_for_synthesis_prompt():
    rendered = _format_selected_items([
        {
            "section": "Sheet: Required Calculations",
            "summary": "Represent source-backed entry e1 in model.xlsx as workbook_row.",
            "entry_id": "e1",
            "target_file": "model.xlsx",
            "native_form": "workbook_row",
            "artifact_function": "workbook_calculation",
            "source": "artifact_commitment",
            "required_source_refs": [{
                "document": "schedule.xlsx",
                "section": "A",
                "evidence": "$1,849,900",
            }],
            "satisfaction_conditions": [
                "Place entry e1 in model.xlsx as a workbook row or table line, not as prose.",
                "Show the calculation expression and final result in separate workbook cells or columns.",
            ],
            "verification_terms": ["$1,849,900", "Net equity"],
        }
    ])

    assert "Artifact-native contract: target=model.xlsx, native=workbook_row, function=workbook_calculation" in rendered
    assert "Verification terms: $1,849,900; Net equity" in rendered
    assert "Required source refs: schedule.xlsx / A: $1,849,900" in rendered
    assert "Satisfaction conditions:" in rendered
    assert "workbook row or table line" in rendered


def test_selected_item_summary_compacts_artifact_wrapper_and_long_text():
    summary = (
        "Represent source-backed entry e1 in memo.docx as memo_statement: "
        + "Specific sourced conclusion. "
        + ("Additional explanatory material. " * 80)
    )

    compact = _compact_selected_item_summary(summary, max_chars=120)

    assert compact.startswith("Specific sourced conclusion.")
    assert "Represent source-backed entry" not in compact
    assert len(compact) <= 123
    assert compact.endswith("...")


def test_selected_evidence_prioritizes_selected_entry_over_early_entries():
    entries = [
        Entry(id=f"e{i}", type="observation", content=f"Filler entry {i} with enough content for rendering.")
        for i in range(1, 90)
    ]
    entries.append(
        Entry(
            id="e250",
            type="calculation",
            content="TAIL_SELECTED_EVIDENCE with exact calculation support.",
        )
    )

    evidence = _selected_evidence_text(
        [{"summary": "Use late evidence.", "entry_id": "e250"}],
        entries,
        max_chars=2000,
    )

    assert "TAIL_SELECTED_EVIDENCE" in evidence
    assert "SELECTED ITEM SUPPORTING ENTRIES" in evidence


def test_file_deliverable_uses_sectioned_drafting_for_large_item_sets():
    blackboard = Blackboard(task_instruction="Prepare a large workbook.")
    blackboard.entries = [
        Entry(id="e1", type="calculation", content="SELECTED_EVIDENCE_SENTINEL supports value 1."),
        Entry(id="e999", type="observation", content="UNRELATED_EVIDENCE_SENTINEL should stay out."),
    ]
    must_include = [
        {"section": "Calculations", "summary": f"Show value {i}.", "entry_id": f"e{i}"}
        for i in range(1, 42)
    ]
    caller = FakeCaller([
        "# Sheet: Calculations\n| Item | Value |\n| --- | --- |\n| SECTIONED_SENTINEL | 1 |",
        "# Sheet: Calculations Part 2\n| Item | Value |\n| --- | --- |\n| SECTIONED_TAIL_SENTINEL | 41 |",
        "# Sheet: Workbook Components\n| Item | Value |\n| WORKBOOK_CRITERION_SENTINEL | present |",
        '{"missing":[],"present_count":41,"missing_count":0}',
    ])

    outputs, _ = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"model": "model.xlsx"},
        [{"id": "C1", "title": "Workbook", "match_criteria": "Workbook", "deliverables": ["model.xlsx"]}],
        caller,
    )

    assert "SECTIONED_SENTINEL" in outputs["model.xlsx"]
    assert "SECTIONED_TAIL_SENTINEL" in outputs["model.xlsx"]
    assert "WORKBOOK_CRITERION_SENTINEL" in outputs["model.xlsx"]
    section_prompts = [p for p in caller.prompts if "Write one section or sheet" in p]
    assert len(section_prompts) == 3
    assert not any("Plan the contents" in p for p in caller.prompts)
    assert "Show value 25." in section_prompts[0]
    assert "Show value 26." not in section_prompts[0]
    assert "Show value 26." in section_prompts[1]
    assert "Show value 41." in section_prompts[1]
    assert "File-specific requirement for model.xlsx: C1" in section_prompts[2]
    assert "SELECTED_EVIDENCE_SENTINEL" in section_prompts[0]
    assert "UNRELATED_EVIDENCE_SENTINEL" not in section_prompts[0]
    assert caller.max_tokens[:3] == [SECTION_DRAFT_MAX_TOKENS] * 3


def test_file_deliverable_chunks_large_single_section_and_tracks_aggregate_usage():
    blackboard = Blackboard(task_instruction="Prepare a very large workbook.")
    must_include = [
        {"section": "Calculations", "summary": f"Show value {i}.", "entry_id": f"e{i}"}
        for i in range(1, 86)
    ]
    caller = FakeCaller([
        "# Sheet: Calculations Part 1\n| Item | Value |\n| --- | --- |\n| CHUNK_1 | 1 |",
        "# Sheet: Calculations Part 2\n| Item | Value |\n| --- | --- |\n| CHUNK_2 | 2 |",
        "# Sheet: Calculations Part 3\n| Item | Value |\n| --- | --- |\n| CHUNK_3 | 3 |",
        "# Sheet: Calculations Part 4\n| Item | Value |\n| --- | --- |\n| CHUNK_4 | 4 |",
        "# Sheet: Workbook Components\n| Item | Value |\n| CRITERION_CHUNK | present |",
        '{"missing":[],"present_count":85,"missing_count":0}',
    ])

    outputs, tokens = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"model": "model.xlsx"},
        [{"id": "C1", "title": "Workbook", "match_criteria": "Workbook", "deliverables": ["model.xlsx"]}],
        caller,
    )

    assert tokens == 90
    assert "CHUNK_1" in outputs["model.xlsx"]
    assert "CHUNK_4" in outputs["model.xlsx"]
    assert "CRITERION_CHUNK" in outputs["model.xlsx"]
    section_prompts = [p for p in caller.prompts if "Write one section or sheet" in p]
    assert len(section_prompts) == 5
    assert "SECTION OR SHEET:\nCalculations Part 1" in section_prompts[0]
    assert "Show value 25." in section_prompts[0]
    assert "Show value 26." not in section_prompts[0]
    assert "SECTION OR SHEET:\nCalculations Part 2" in section_prompts[1]
    assert "Show value 26." in section_prompts[1]
    assert "Show value 50." in section_prompts[1]
    assert "Show value 51." not in section_prompts[1]
    assert "SECTION OR SHEET:\nCalculations Part 3" in section_prompts[2]
    assert "Show value 51." in section_prompts[2]
    assert "Show value 75." in section_prompts[2]
    assert "SECTION OR SHEET:\nCalculations Part 4" in section_prompts[3]
    assert "Show value 76." in section_prompts[3]
    assert "Show value 85." in section_prompts[3]
    assert "SECTION OR SHEET:\nWorkbook Components" in section_prompts[4]
    assert "File-specific requirement for model.xlsx: C1" in section_prompts[4]
    assert caller.max_tokens[:5] == [SECTION_DRAFT_MAX_TOKENS] * 5

    blackboard.add_tokens_from_last_call(tokens)

    assert blackboard.total_tokens_used == 90
    assert blackboard.tokens_input == 60
    assert blackboard.tokens_output == 30
    assert blackboard.cost_by_model["fake-model"] == {
        "input": 60,
        "output": 30,
        "total": 90,
        "calls": 6,
    }


def test_completeness_verifier_keeps_tail_context_for_long_drafts():
    blackboard = Blackboard(task_instruction="Prepare a long deliverable.")
    draft = "HEAD_CONTEXT\n" + ("x" * 170000) + "\nTAIL_CONTEXT"
    caller = FakeCaller(['{"missing":[],"present_count":1,"missing_count":0}'])

    _verify_completeness(
        draft,
        [{"summary": "Tail fact", "entry_id": "e1"}],
        blackboard,
        caller,
    )

    prompt = caller.prompts[0]
    assert "HEAD_CONTEXT" in prompt
    assert "TAIL_CONTEXT" in prompt
    assert "tail of long draft follows" in prompt


def test_criteria_only_deliverables_are_not_used_for_generation():
    task_data = {
        "criteria": [
            {"deliverables": ["memo.docx"]},
            {"deliverables": ["model.xlsx", "memo.docx"]},
        ]
    }

    assert _deliverables_for_task(task_data) == {}


def test_sectioned_file_deliverable_includes_tail_criteria():
    blackboard = Blackboard(task_instruction="Prepare a large memo.")
    must_include = [
        {"section": "Analysis", "summary": f"Discuss item {i}.", "entry_id": f"e{i}"}
        for i in range(1, 42)
    ]
    criteria = [
        {
            "id": f"C{i}",
            "title": f"Criterion {i}",
            "match_criteria": f"Match detail {i}",
            "deliverables": ["memo.docx"],
        }
        for i in range(1, 102)
    ]
    caller = FakeCaller([
        "SECTION_ONE",
        "SECTION_TWO",
        "CRITERIA_ONE",
        "CRITERIA_TWO",
        "CRITERIA_THREE",
        "CRITERIA_FOUR",
        "CRITERIA_FIVE",
        '{"missing":[],"present_count":41,"missing_count":0}',
    ])

    outputs, _ = synthesize_file_deliverables(
        blackboard,
        must_include,
        {"memo": "memo.docx"},
        criteria,
        caller,
    )

    assert "SECTION_ONE" in outputs["memo.docx"]
    assert "CRITERIA_FIVE" in outputs["memo.docx"]
    section_prompts = [p for p in caller.prompts if "Write one section or sheet" in p]
    assert len(section_prompts) == 7
    assert "C101" in section_prompts[-1]
    assert "Match detail 101" in section_prompts[-1]
    assert "additional criteria omitted" not in section_prompts[-1]


def test_file_scoped_synthesis_routing_keeps_ordinary_single_docx_generic():
    assert _should_use_file_scoped_synthesis({
        "manual": "compliance-manual.docx",
    }) is False

    assert _should_use_file_scoped_synthesis({
        "analysis": "stipulation-markup-analysis.docx",
    }) is False

    assert _should_use_file_scoped_synthesis({
        "schedule": "whitford-asset-schedule.docx",
    }) is False

    assert _should_use_file_scoped_synthesis({
        "memo": "memo.docx",
        "model": "model.xlsx",
    }) is True

    assert _should_use_file_scoped_synthesis({
        "model": "obligation-tracker.xlsx",
    }) is True

    assert _should_use_file_scoped_synthesis({
        "redline": "agreement-redline.docx",
    }) is True

    assert _should_use_file_scoped_synthesis({
        "markup": "credit-agreement-markup.docx",
    }) is True

    assert _should_use_file_scoped_synthesis({
        "rider": "lease-rider.docx",
    }) is True
